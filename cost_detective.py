from datetime import datetime
import boto3
from openpyxl import Workbook, load_workbook
import requests
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.protection import Protection
import calendar
from operator import itemgetter
from msal import ConfidentialClientApplication
from dotenv import load_dotenv
import os
def get_month_dates(month):
    current_year = datetime.now().year
     
    # Convert month name to its corresponding number
    month_number = list(calendar.month_name).index(month.capitalize())
    
    # Get the number of days in the month
    _, num_days = calendar.monthrange(2024, month_number)
    
    # Construct start and end dates for the current month
    start_date_current = f"{current_year}-{month_number:02d}-01"
    end_date_current = f"{current_year}-{month_number:02d}-{num_days}"
    
    # Calculate start and end dates for the previous month
    if month_number == 1:
        prev_month_number = 12
        prev_year = current_year-1
    else:
        prev_month_number = month_number - 1
        prev_year = current_year
    
    _, prev_num_days = calendar.monthrange(prev_year, prev_month_number)
    start_date_previous = f"{prev_year}-{prev_month_number:02d}-01"
    end_date_previous = f"{prev_year}-{prev_month_number:02d}-{prev_num_days}"
    
    return (start_date_previous, end_date_previous), (start_date_current, end_date_current)


def accounts_access(previous_months, current_months, ws):
    
    ce_client = boto3.client('ce')
    org_client = boto3.client('organizations')
    account_paginator = org_client.get_paginator('list_accounts')
    account_iterator = account_paginator.paginate()
    for accounts in account_iterator:
        for account in accounts['Accounts']:
            account_id = account['Id']
            print(account['Id'])
            previous_costs = analyse_costs(ce_client, previous_months[0], previous_months[1], account_id)
            current_costs = analyse_costs(ce_client, current_months[0], current_months[1], account_id)
            cost_per_account_per_month(previous_costs, current_costs, account['Name'], ws)
            
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 1000:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
    for row in ws.iter_rows(min_row=2, max_row=100 + 1, min_col=3, max_col=16):
        for cell in row[2::3]:  # Select every third cell starting from the third one (cost difference cells)
            cell.number_format = '$#,##0.00'  # Currency format


def cost_per_account_per_month(previous_costs, current_costs, account, ws):
        # Calculate cost difference for each service
    cost_differences = {}
    for service, current_cost in current_costs.items():
        if service in previous_costs:
            previous_cost = previous_costs[service]
            cost_differences[service] = current_cost - previous_cost

    # Find top 3 services with the greatest cost difference
    top_services = sorted(cost_differences.items(), key=lambda x: x[1], reverse=True)[:5]
    row_data = [account]
    for service, cost_difference in top_services:
        row_data.extend([service, cost_difference, ""])

    # Fill in empty cells if less than 3 services are found
    if len(top_services) < 5:
        row_data.extend([''] * (5 - len(top_services)) * 3)

    # Write data row
    ws.append(row_data)
    if top_services:
    # Print the results
        print(f'{account} Top 3 services with the greatest cost difference:')
        for service, cost_difference in top_services:
            print(f"{service}: ${cost_difference:.2f}")
    
def analyse_costs(ce_client, start, end, account_id):
    # Analyze costs associated with ECS for a specific account
    response = ce_client.get_cost_and_usage(
        TimePeriod={'Start': start, 'End': end},
        Granularity='MONTHLY',
        Metrics=['UnblendedCost'],  # Use AmortizedCost to include all costs, as this is what stax use for data.
        GroupBy=[{'Type': 'DIMENSION', 'Key': 'SERVICE'}],# Group by service
        Filter= {
            'And': [ {'And':[
                {'Dimensions': {'Key': 'BILLING_ENTITY', 'Values': ['AWS']}},
                {'Dimensions': {'Key': 'LINKED_ACCOUNT', 'Values': [account_id]}}]},
                    
                {'Not': {'Dimensions': {'Key': 'SERVICE', 'Values': ['Tax']}}}
            ]
        },
    )
       # Extract costs for each service
    service_costs = {}
    for result in response['ResultsByTime']:
        for group in result['Groups']:
            service_name = group['Keys'][0]
            cost = float(group['Metrics']['UnblendedCost']['Amount'])
            service_costs[service_name] = cost
    return service_costs
         

#script to get the costs for the entire team 
###################################################################################################################

def team_accounts_access(previous_months, current_months, team_accounts, month_year, ws):
    

    ce_client = boto3.client('ce')
    org_client = boto3.client('organizations')
    account_paginator = org_client.get_paginator('list_accounts')
    account_iterator = account_paginator.paginate()
    team_costs = {'team1': {} , 'team2': {}, 'team3': {}}
    for accounts in account_iterator:
        for account in accounts['Accounts']:
            account_id = account['Id']
            print(account['Id'])
            previous_costs = analyse_costs(ce_client, previous_months[0], previous_months[1], account_id)
            previous_costs_values = [value for value in previous_costs.values() if isinstance(value, (int, float))]
            previous_costs_sum = sum(previous_costs_values)
            current_costs = analyse_costs(ce_client, current_months[0], current_months[1], account_id)
            current_costs_values = [value for value in current_costs.values() if isinstance(value, (int, float))]
            current_costs_sum = sum(current_costs_values)
            difference_costs = current_costs_sum - previous_costs_sum
            
            for team in team_accounts.keys():
                if account['Name'] in team_accounts[team]:
                    team_costs[team][account['Name']] = difference_costs
    
    team_costs_list = [{key: value} for key, value in team_costs.items()]

    print(team_costs_list)
    return team_costs_list, month_year

def get_bar_graph(team_costs, month_year, ws):

    chart_row = 152  # Start row for the first chart
    insert_row = 2
    data_row = 150
    top_n = 5  # Number of top values to display
    
    for idx, data in enumerate(team_costs, start=1):
        group, subgroup_data = next(iter(data.items()))
        
        sorted_data = sorted(subgroup_data.items(), key=itemgetter(1), reverse=True)[:top_n]
        top_subgroup_data = dict(sorted_data)
        # Write data to the worksheet
        ws.cell(row=data_row, column=1, value='Accounts')
        ws.append([f'DataSet {idx}'])
        ws.append(['Groups', *top_subgroup_data.keys(), ''])
        ws.append(['Accounts', *top_subgroup_data.values(), ''])

        # Create a bar chart
        chart = BarChart()
        chart.type = "col"
        chart.title = f'{group} Data {month_year}'
        chart.y_axis.title = "Cost Difference ($)"
        chart.y_axis.number_format = '$#,##0.00'

        data_ref = Reference(ws, min_col=2, min_row=chart_row, max_col=len(top_subgroup_data) + 1, max_row=chart_row + 1)
        cats = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.x_axis.tickLblPos = 'low'
        # Add the chart to the worksheet
        chart_row += 3
        data_row += 3
        ws.add_chart(chart, f"U{insert_row}")
        insert_row += 15

            


def download_file_sharepoint(tenant_id, client_id, client_secret, site_url):
    
    authority = f'https://login.microsoftonline.com/{tenant_id}'

    app = ConfidentialClientApplication(
        client_id,
        authority=authority,
        client_credential=client_secret,
    )

    token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    access_token = token_response.get("access_token")
    headers = {
        "Authorization": "Bearer " + access_token,
        "Content-Type": "application/json"
    }

    site_url = "https://graph.microsoft.com/v1.0/sites//yourcompany.sharepoint.com:/teams/folder"

    response = requests.get(site_url, headers=headers)
    site_id = response.json().get('id')

    drivers_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"

    headers = {
        "Authorization": "Bearer " + access_token
    }
    response = requests.get(drivers_url, headers=headers)
    drives = response.json()
    drive_id = drives['value'][0]['id']
    folder_path = f"YourFolder"

    items_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{folder_path}"
    response = requests.get(items_url, headers=headers)
    file_name = 'cost_detective_report1.xlsx'
    file_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{folder_path}/{file_name}:/content"

    response = requests.get(file_url, headers=headers, allow_redirects=True)

    if response.status_code == 200:
        with open(file_name, 'wb') as f:
            f.write(response.content)
        print(f"File downloaded successfully: {file_name}")
    else:
        print("Failed to download file:", response.json())
    return drive_id, folder_path, file_name, access_token

   

def upload_new_month_file(drive_id,folder_path,file_name, access_token):
        
    # Endpoint to upload file

    upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{folder_path}/{file_name}:/content"

    # Set up headers
    headers = {
        "Authorization": "Bearer " + access_token,
        "Content-Type": "application/octet-stream"
    }
    # Read the file content
    with open(file_name, 'rb') as f:
        file_content = f.read()
    # Upload the file
    response = requests.put(upload_url, headers=headers, data=file_content)

    # Check the response
    if response.status_code == 200:
        print("File uploaded successfully.")
    else:
        print("Failed to upload file:", response.json())


def create_workbook(month):
    tenant_id = os.getenv('TENANT_ID')
    client_id = os.getenv('CLIENT_ID')
    client_secret = os.getenv('CLIENT_SECRET')
    site_url = 'https://yourcompany.sharepoint.com/teams/folder'

    drive_id, folder_path, file_name, access_token = download_file_sharepoint(tenant_id, client_id, client_secret, site_url)
    try:
        wb = load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()


    previous_month_dates, current_month_dates = get_month_dates(month)
    print(previous_month_dates)

    month_year = month + '24'

    team_accounts = {
        'Blanks': ['aws1' 'aws2', 'aws3'],
        'Retail': ['aws4', 'aws5', 'aws6'],}

    ws = wb.create_sheet(title=f'{month_year}')

    # Write header row
    header_row = (["Account", "Top Service 1", "Cost Difference 1", "Justification 1", 
            "Top Service 2", "Cost Difference 2", "Justification 2",
            "Top Service 3", "Cost Difference 3", "Justification 3", 
            "Top Service 4", "Cost Difference 4",  "Justification 4", 
            "Top Service 5", "Cost Difference 5", "Justification 5"])

    ws.append(header_row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey fill

    ws.auto_filter.ref = ws.dimensions

    return wb, ws, previous_month_dates, current_month_dates, team_accounts, month_year, drive_id, folder_path, access_token

def edit_save_workbook(ws, file_name, wb):
    ws.protection.sheet = True
    ws.protection.password = "cloud"  # Set a password for unlocking the sheet

    # Unlock specific columns: D, G, J, M, and P
    unlocked_columns = [4, 7, 10, 13, 16]  # Columns D, G, J, M, and P
    for col in unlocked_columns:
        for row in ws.iter_rows(min_row=1, max_row=10, min_col=col, max_col=col):
            for cell in row:
                cell.protection = Protection(locked=False)  # Unlock the cells

    wb.save(file_name)
    
    
    


def main():
    #The datetime for the month is used for the Lambda, but just hardcoded it here for testing purposes
    #month = datetime.now().strftime('%B')
    month = 'March'
    file_name = 'cost_detective_report1.xlsx' 
    wb, ws, previous_month_dates, current_month_dates, team_accounts, month_year, drive_id, folder_path, access_token = create_workbook(month)
        
    accounts_access(previous_month_dates, current_month_dates, ws)
    team_costs_list, month_year = team_accounts_access(previous_month_dates, current_month_dates, team_accounts, month_year, ws)
    get_bar_graph(team_costs_list, month_year, ws)

    edit_save_workbook(ws, file_name, wb)
    upload_new_month_file(drive_id,folder_path,file_name, access_token)

if __name__ == "__main__":
    load_dotenv()
    main()
    
    